#!/usr/bin/env python3
"""
Excel Exporter for Playlist Indexer
Exports playlist video data to Excel format with Name, Link, Description columns.

Usage:
    # Export single playlist
    python excel_exporter.py --playlist daisy_synths_to_check --output export.xlsx
    
    # Export all playlists
    python excel_exporter.py --all --output all_playlists.xlsx
"""

import os
import json
import argparse
from typing import List, Dict, Optional
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    exit(1)


def load_playlist_registry(output_dir: str = "output") -> Dict:
    """Load the playlists registry."""
    registry_path = os.path.join(output_dir, "playlists.json")
    if not os.path.exists(registry_path):
        raise FileNotFoundError(f"Registry not found: {registry_path}")
    
    with open(registry_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_playlist_videos(playlist_id: str, output_dir: str = "output") -> List[Dict]:
    """Load video data for a specific playlist."""
    registry = load_playlist_registry(output_dir)
    
    # Find playlist in registry
    playlist_info = None
    for p in registry['playlists']:
        if p['id'] == playlist_id:
            playlist_info = p
            break
    
    if not playlist_info:
        raise ValueError(f"Playlist not found: {playlist_id}")
    
    # Load video data
    data_file = os.path.join(playlist_info['output_dir'], f"{playlist_id}_data.json")
    if not os.path.exists(data_file):
        raise FileNotFoundError(f"Video data not found: {data_file}")
    
    with open(data_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_all_videos(output_dir: str = "output") -> List[Dict]:
    """Load all videos from all playlists."""
    registry = load_playlist_registry(output_dir)
    all_videos = []
    
    for playlist in registry['playlists']:
        data_file = os.path.join(playlist['output_dir'], f"{playlist['id']}_data.json")
        if os.path.exists(data_file):
            with open(data_file, 'r', encoding='utf-8') as f:
                videos = json.load(f)
                all_videos.extend(videos)
    
    return all_videos


def export_to_excel(videos: List[Dict], output_path, playlist_name: str = "All Playlists"):
    """
    Export videos to Excel with Name, Link, Description columns.
    
    Args:
        videos: List of video dictionaries
        output_path: Path string or file-like object (e.g., BytesIO) for the output
        playlist_name: Name to show in the header
        
    Returns:
        The output_path (unchanged)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Videos"
    
    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["#", "Name", "Link", "Description", "Playlist", "Thematic", "Genre", "Length", "Author", "Tags"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for idx, video in enumerate(videos, 1):
        row = idx + 1
        
        # Index
        ws.cell(row=row, column=1, value=idx).border = thin_border
        
        # Name (Title)
        name_cell = ws.cell(row=row, column=2, value=video.get('title', 'Unknown'))
        name_cell.alignment = cell_alignment
        name_cell.border = thin_border
        
        # Link (URL as hyperlink)
        url = video.get('url', '')
        link_cell = ws.cell(row=row, column=3, value=url)
        if url:
            link_cell.hyperlink = url
            link_cell.font = link_font
        link_cell.alignment = cell_alignment
        link_cell.border = thin_border
        
        # Description (truncate if too long)
        description = video.get('description', '')
        if len(description) > 500:
            description = description[:497] + "..."
        desc_cell = ws.cell(row=row, column=4, value=description)
        desc_cell.alignment = cell_alignment
        desc_cell.border = thin_border
        
        # Playlist name
        playlist_cell = ws.cell(row=row, column=5, value=video.get('playlist_name', ''))
        playlist_cell.alignment = cell_alignment
        playlist_cell.border = thin_border
        
        # Metadata access helper
        metadata = video.get('metadata', {})
        if not metadata and 'thematic' not in video:
             # Legacy or flat structure support if needed/fallback
             pass
             
        # Thematic
        thematic = "Unknown"
        if isinstance(metadata, dict):
             thematic_data = metadata.get('thematic', {})
             if isinstance(thematic_data, dict):
                 thematic = thematic_data.get('primary', 'Unknown')
             elif isinstance(thematic_data, str): # Legacy/Migration edge case
                 thematic = thematic_data
        
        ws.cell(row=row, column=6, value=thematic).border = thin_border
        
        # Genre
        genre = "Unknown"
        if isinstance(metadata, dict):
             genre_data = metadata.get('genre', {})
             if isinstance(genre_data, dict):
                 genre = genre_data.get('primary', 'Unknown')
        
        ws.cell(row=row, column=7, value=genre).border = thin_border
        
        # Length Category
        length = "Unknown"
        if isinstance(metadata, dict):
            length = metadata.get('length_category', 'Unknown')
        ws.cell(row=row, column=8, value=length).border = thin_border
        
        # Author Type
        author_type = "Unknown"
        if isinstance(metadata, dict):
            author_type = video.get('author_type', 'Unknown') # Check if it's on root or metadata
            # Plan says metadata.author_type
            author_type = metadata.get('author_type', 'Unknown')
        ws.cell(row=row, column=9, value=author_type).border = thin_border
        
        # Tags (Combined)
        tags_raw = video.get('tags', [])
        tags_str = ""
        if isinstance(tags_raw, dict):
            tags_str = ", ".join(tags_raw.get('combined', []))
        elif isinstance(tags_raw, list):
            tags_str = ", ".join(tags_raw)
            
        ws.cell(row=row, column=10, value=tags_str).border = thin_border
    
    # Column widths
    ws.column_dimensions['A'].width = 5   # Index
    ws.column_dimensions['B'].width = 50  # Name
    ws.column_dimensions['C'].width = 45  # Link
    ws.column_dimensions['D'].width = 60  # Description - reduced slightly
    ws.column_dimensions['E'].width = 25  # Playlist
    ws.column_dimensions['F'].width = 20  # Thematic
    ws.column_dimensions['G'].width = 15  # Genre
    ws.column_dimensions['H'].width = 12  # Length
    ws.column_dimensions['I'].width = 15  # Author
    ws.column_dimensions['J'].width = 40  # Tags
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = f"A1:J{len(videos) + 1}"
    
    # Save
    wb.save(output_path)
    return output_path


def export_playlist(playlist_id: str, output_path: str, output_dir: str = "output") -> str:
    """Export a single playlist to Excel."""
    videos = load_playlist_videos(playlist_id, output_dir)
    return export_to_excel(videos, output_path, playlist_id)


def export_all_playlists(output_path: str, output_dir: str = "output") -> str:
    """Export all playlists to a single Excel file."""
    videos = load_all_videos(output_dir)
    return export_to_excel(videos, output_path, "All Playlists")


def main():
    """CLI entry point."""
    parser = argparse.ArgumentParser(description="Export playlist videos to Excel")
    parser.add_argument('--playlist', '-p', help='Playlist ID to export')
    parser.add_argument('--all', '-a', action='store_true', help='Export all playlists')
    parser.add_argument('--output', '-o', required=True, help='Output Excel file path')
    parser.add_argument('--output-dir', default='output', help='Directory with indexed playlists')
    
    args = parser.parse_args()
    
    if not args.playlist and not args.all:
        print("Error: Specify --playlist ID or --all")
        return 1
    
    try:
        if args.all:
            result = export_all_playlists(args.output, args.output_dir)
            print(f"✅ Exported all playlists to: {result}")
        else:
            result = export_playlist(args.playlist, args.output, args.output_dir)
            print(f"✅ Exported {args.playlist} to: {result}")
        return 0
    except Exception as e:
        print(f"❌ Error: {e}")
        return 1


if __name__ == "__main__":
    exit(main())
